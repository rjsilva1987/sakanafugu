import os
import io
import re
import json
import base64
import zipfile
import mimetypes
import chardet

from flask import Flask, request, jsonify, render_template, Response, stream_with_context
from openai import OpenAI
from dotenv import load_dotenv

# ── Firebase ───────────────────────────────────────────────────────────────
import firebase_admin
from firebase_admin import credentials, storage

load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB

# ── Env ────────────────────────────────────────────────────────────────────
SAKANA_API_KEY     = os.getenv("SAKANA_API_KEY", "")
MODEL              = os.getenv("MODEL", "fugu")
REASONING_EFFORT   = os.getenv("REASONING_EFFORT", "high")
FIREBASE_CREDS     = os.getenv("FIREBASE_CREDENTIALS", "firebase-credentials.json")
FIREBASE_BUCKET    = os.getenv("FIREBASE_STORAGE_BUCKET", "")

# ── Firebase init ──────────────────────────────────────────────────────────
firebase_ok = False
try:
    cred = credentials.Certificate(FIREBASE_CREDS)
    firebase_admin.initialize_app(cred, {"storageBucket": FIREBASE_BUCKET})
    firebase_ok = True
except Exception as e:
    print(f"⚠️  Firebase não inicializado: {e}")

# ── Sakana client ──────────────────────────────────────────────────────────
client = OpenAI(
    base_url="https://api.sakana.ai/v1",
    api_key=SAKANA_API_KEY,
)

SYSTEM_PROMPT = (
    "Você é um assistente inteligente alimentado pelo Sakana Fugu. "
    "Responda sempre no mesmo idioma que o usuário usar. "
    "Seja claro, direto e útil. "
    "Quando receber arquivos, analise o conteúdo com atenção antes de responder. "
    "Arquivos são apresentados entre marcadores [ARQUIVO: nome] e [/ARQUIVO]."
)

IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp"}

# Extensões que podemos tentar ler como texto
TEXT_EXTENSIONS = {
    ".txt", ".md", ".csv", ".tsv", ".json", ".jsonl", ".xml", ".html",
    ".htm", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".env", ".log",
    ".py", ".js", ".ts", ".jsx", ".tsx", ".css", ".scss", ".sh", ".bat",
    ".sql", ".r", ".rb", ".go", ".rs", ".java", ".c", ".cpp", ".h",
    ".cs", ".php", ".swift", ".kt", ".dart", ".vue", ".svelte",
}


# ══════════════════════════════════════════════════════════════════════════
# Extração de conteúdo
# ══════════════════════════════════════════════════════════════════════════

def decode_bytes(raw: bytes, filename: str = "") -> str | None:
    """Tenta decodificar bytes como texto. Retorna None se for binário."""
    ext = os.path.splitext(filename)[1].lower()
    if ext and ext not in TEXT_EXTENSIONS and ext not in {".pdf", ".docx", ".xlsx"}:
        # Extensão conhecida como binária — não tenta
        return None
    detected = chardet.detect(raw[:4096])
    enc = detected.get("encoding") or "utf-8"
    try:
        return raw.decode(enc, errors="replace")
    except Exception:
        return None


def extract_pdf(raw: bytes, filename: str) -> str:
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(stream=raw, filetype="pdf")
        pages = []
        for i, page in enumerate(doc):
            text = page.get_text()
            if text.strip():
                pages.append(f"[Página {i+1}]\n{text.strip()}")
        return "\n\n".join(pages) if pages else "(PDF sem texto extraível)"
    except Exception as e:
        return f"(Erro ao ler PDF: {e})"


def extract_docx(raw: bytes, filename: str) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(raw))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs) or "(DOCX sem texto)"
    except Exception as e:
        return f"(Erro ao ler DOCX: {e})"


def extract_xlsx(raw: bytes, filename: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append("\t".join(cells))
            if rows:
                parts.append(f"[Aba: {sheet}]\n" + "\n".join(rows))
        return "\n\n".join(parts) or "(Planilha vazia)"
    except Exception as e:
        return f"(Erro ao ler XLSX: {e})"


def extract_file_content(raw: bytes, filename: str) -> dict:
    """
    Retorna dict com:
      - type: 'text' | 'image' | 'binary'
      - content: str (texto extraído) ou bytes (imagem)
      - mime: str
    """
    ext  = os.path.splitext(filename)[1].lower()
    mime = mimetypes.guess_type(filename)[0] or "application/octet-stream"

    # Imagem
    if mime in IMAGE_TYPES:
        return {"type": "image", "content": raw, "mime": mime}

    # PDF
    if ext == ".pdf" or mime == "application/pdf":
        return {"type": "text", "content": extract_pdf(raw, filename), "mime": mime}

    # DOCX
    if ext == ".docx":
        return {"type": "text", "content": extract_docx(raw, filename), "mime": mime}

    # XLSX
    if ext in {".xlsx", ".xlsm"}:
        return {"type": "text", "content": extract_xlsx(raw, filename), "mime": mime}

    # Texto genérico
    text = decode_bytes(raw, filename)
    if text is not None:
        return {"type": "text", "content": text, "mime": mime}

    # Binário desconhecido — avisa mas não trava
    return {
        "type": "binary",
        "content": f"(Arquivo binário: {filename}, {len(raw)} bytes — conteúdo não extraível)",
        "mime": mime,
    }


def extract_zip(raw: bytes) -> list[dict]:
    """Extrai todos os arquivos de um ZIP e retorna lista de {name, extracted}."""
    results = []
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                # Ignora arquivos ocultos e __MACOSX
                name = info.filename
                if any(part.startswith((".", "__MACOSX")) for part in name.split("/")):
                    continue
                try:
                    file_raw = zf.read(name)
                    extracted = extract_file_content(file_raw, os.path.basename(name))
                    results.append({"name": name, "extracted": extracted})
                except Exception as e:
                    results.append({
                        "name": name,
                        "extracted": {"type": "text", "content": f"(Erro ao ler: {e})", "mime": ""},
                    })
    except zipfile.BadZipFile:
        results.append({
            "name": "erro",
            "extracted": {"type": "text", "content": "(Arquivo ZIP corrompido ou inválido)", "mime": ""},
        })
    return results


# ══════════════════════════════════════════════════════════════════════════
# Firebase Storage
# ══════════════════════════════════════════════════════════════════════════

def upload_to_firebase(raw: bytes, filename: str, mime: str) -> str | None:
    """Faz upload para Firebase Storage e retorna a URL pública."""
    if not firebase_ok:
        return None
    try:
        bucket = storage.bucket()
        blob   = bucket.blob(f"uploads/{filename}")
        blob.upload_from_string(raw, content_type=mime)
        blob.make_public()
        return blob.public_url
    except Exception as e:
        print(f"Firebase upload error: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════
# Monta content_parts para a API
# ══════════════════════════════════════════════════════════════════════════

def build_content_parts(files, user_text: str) -> list[dict]:
    parts = []

    for f in files:
        raw      = f.read()
        filename = f.filename
        mime     = mimetypes.guess_type(filename)[0] or f.content_type or "application/octet-stream"
        ext      = os.path.splitext(filename)[1].lower()

        # Faz upload no Firebase (qualquer arquivo)
        upload_to_firebase(raw, filename, mime)

        # ZIP — expande e processa cada arquivo interno
        if ext == ".zip" or mime == "application/zip":
            inner_files = extract_zip(raw)
            summary_lines = [f"[ZIP: {filename}] contém {len(inner_files)} arquivo(s):"]
            for item in inner_files:
                ext_inner = extract_file_content  # referência para reúso
                ex = item["extracted"]
                summary_lines.append(f"  • {item['name']} ({ex['mime'] or 'binário'})")

            # Inclui conteúdo de cada arquivo interno
            for item in inner_files:
                ex = item["extracted"]
                if ex["type"] == "image":
                    b64 = base64.standard_b64encode(ex["content"]).decode()
                    parts.append({
                        "type": "image_url",
                        "image_url": {"url": f"data:{ex['mime']};base64,{b64}"},
                    })
                else:
                    text_block = (
                        f"[ARQUIVO: {item['name']}]\n"
                        f"{ex['content'][:8000]}"
                        + (" …(truncado)" if len(ex["content"]) > 8000 else "")
                        + "\n[/ARQUIVO]"
                    )
                    parts.append({"type": "text", "text": text_block})

            # Sumário do ZIP no início
            parts.insert(0, {"type": "text", "text": "\n".join(summary_lines)})
            continue

        # Imagem
        ex = extract_file_content(raw, filename)
        if ex["type"] == "image":
            b64 = base64.standard_b64encode(raw).decode()
            parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{b64}"},
            })
        else:
            content = ex["content"]
            text_block = (
                f"[ARQUIVO: {filename}]\n"
                f"{content[:8000]}"
                + (" …(truncado)" if len(content) > 8000 else "")
                + "\n[/ARQUIVO]"
            )
            parts.append({"type": "text", "text": text_block})

    if user_text:
        parts.append({"type": "text", "text": user_text})

    return parts


# ══════════════════════════════════════════════════════════════════════════
# Rotas
# ══════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html", model=MODEL)


@app.route("/api/chat", methods=["POST"])
def chat():
    if not SAKANA_API_KEY:
        return jsonify({"error": "SAKANA_API_KEY não configurada no arquivo .env"}), 500

    if request.content_type and "multipart/form-data" in request.content_type:
        raw_history = request.form.get("messages", "[]")
        try:
            history = json.loads(raw_history)
        except Exception:
            return jsonify({"error": "Campo 'messages' inválido."}), 400

        user_text = request.form.get("text", "").strip()
        files     = request.files.getlist("files")

        content_parts = build_content_parts(files, user_text)

        if not content_parts:
            return jsonify({"error": "Envie texto ou arquivo."}), 400

        current_msg = {"role": "user", "content": content_parts}
        messages    = history + [current_msg]

    else:
        data     = request.get_json(silent=True) or {}
        messages = data.get("messages", [])
        if not messages:
            return jsonify({"error": "Nenhuma mensagem enviada."}), 400

    full_messages = [{"role": "system", "content": SYSTEM_PROMPT}] + messages

    def generate():
        try:
            stream = client.chat.completions.create(
                model=MODEL,
                messages=full_messages,
                stream=True,
                extra_body={"reasoning": {"effort": REASONING_EFFORT}},
                timeout=120.0,
            )
            for chunk in stream:
                delta = chunk.choices[0].delta
                if delta.content:
                    yield delta.content
        except Exception as e:
            yield f"\n\n[Erro: {str(e)}]"

    return Response(stream_with_context(generate()), mimetype="text/plain")


@app.route("/api/config")
def config():
    return jsonify({
        "model": MODEL,
        "reasoning_effort": REASONING_EFFORT,
        "api_configured": bool(SAKANA_API_KEY),
        "firebase_ok": firebase_ok,
    })


if __name__ == "__main__":
    print("🐡 Sakana Fugu Chatbot iniciado")
    print(f"   Modelo: {MODEL}  |  Esforço: {REASONING_EFFORT}")
    print(f"   Firebase: {'✅ conectado' if firebase_ok else '⚠️  não configurado'}")
    print("   Acesse: http://localhost:5000")
    app.run(debug=True, port=5000)
